#!/usr/bin/env python3
"""
Catalog Standardizer - Convert any catalog format to VAYA standard format
Handles Excel files and PDFs from various suppliers
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import sys
import os
from pathlib import Path

# VAYA Standard Column Names (in order)
VAYA_COLUMNS = [
    'Collection',
    'Design Name', 
    'Composition',
    'Weight/Mt',
    'Design Horizontal',
    'Design Vertical',
    'Fabric Width (cm)',
    'Martindale',
    'End Use',
    'HS Code',
    'Updated DP/Mtr (Cut Rate) May 2025',
    'GST',
    'Dealer Price after GST (Cut Rate)',
    'Dealer Price/Mtr (Roll Rate)',
    'Dealer Price after GST (Roll Rate)',
    'RR Price (Cut Rate)',
    'RR Price after GST (Cut Rate)'
]

# Column mapping for different suppliers
COLUMN_MAPPINGS = {
    'SANSAAR': {
        'COLLECTION': 'Collection',
        'Description': 'Design Name',
        'Material Code': 'HS Code',
        'CL_RATE': 'Updated DP/Mtr (Cut Rate) May 2025',
        'GST': 'GST',
        'RRP With GST': 'RR Price after GST (Cut Rate)'
    },
    'FF_A_dress': {
        'Collection name': 'Collection',
        'Design': 'Design Name',
        'HSN Codes': 'HS Code',
        ' Tax %': 'GST',
        'NEW DP': 'Updated DP/Mtr (Cut Rate) May 2025',
        'NEW RC': 'RR Price (Cut Rate)',
        'NEW MRP': 'RR Price after GST (Cut Rate)'
    },
    'FABRIZIO': {
        'COLLECTION NAME': 'Collection',
        'PRODUCT NAME': 'Design Name',
        'COMPOSITION': 'Composition',
        'MARTINDALE': 'Martindale',
        'HSNCODE': 'HS Code',
        'WIDTH (INCM)': 'Fabric Width (cm)',
        'WIDTH(INCM)': 'Fabric Width (cm)',
        'GSM': 'Weight/Mt',
        'GST': 'GST',
        'DP': 'Updated DP/Mtr (Cut Rate) May 2025'
    }
}

def clean_numeric_value(value):
    """Clean and convert numeric values"""
    if pd.isna(value):
        return None
    
    if isinstance(value, (int, float)):
        return value
    
    # Remove currency symbols, commas, percentage signs
    if isinstance(value, str):
        value = value.strip().replace('₹', '').replace('/-', '').replace(',', '').replace('%', '')
        try:
            return float(value)
        except:
            return None
    return None

def detect_catalog_type(df, filename):
    """Improved detection to prevent FF_A_dress from being identified as FABRIZIO"""
    filename_upper = filename.upper()
    columns_str = '|'.join([str(c).upper() for c in df.columns])
    
    # 1. Check Filename FIRST (Highest Priority)
    if 'SANSAAR' in filename_upper:
        return 'SANSAAR'
    if 'FF_A_DRESS' in filename_upper or 'DIVINE' in filename_upper or 'F&F' in filename_upper:
        return 'FF_A_dress'
    if 'FABRIZIO' in filename_upper:
        return 'FABRIZIO'
    
    # 2. Specific Column Keyword Detection
    # FF_A_dress specific unique columns
    if 'NEW DP' in columns_str or 'NEW MRP' in columns_str:
        return 'FF_A_dress'
    
    # Fabrizio specific (be more specific than just HSN)
    if 'PRODUCT NAME' in columns_str and 'HSNCODE' in columns_str:
        return 'FABRIZIO'
        
    if 'COLLECTION' in columns_str and 'SERIALNO' in columns_str:
        return 'SANSAAR'
    
    return 'UNKNOWN'

def map_columns(df, catalog_type):
    """Map columns from source format to VAYA standard"""
    if catalog_type not in COLUMN_MAPPINGS:
        return df
    
    mapping = COLUMN_MAPPINGS[catalog_type]
    
    # Create a new dataframe with VAYA columns
    standardized_df = pd.DataFrame()
    
    # Map available columns
    for source_col, target_col in mapping.items():
        if source_col in df.columns:
            standardized_df[target_col] = df[source_col]
    
    # Add missing columns with empty values
    for col in VAYA_COLUMNS:
        if col not in standardized_df.columns:
            standardized_df[col] = None
    
    # Reorder to match VAYA column order
    standardized_df = standardized_df[VAYA_COLUMNS]
    
    return standardized_df

def calculate_gst_prices(df):
    """Calculate GST-inclusive prices if missing"""
    
    # Clean numeric columns
    if 'Updated DP/Mtr (Cut Rate) May 2025' in df.columns:
        df['Updated DP/Mtr (Cut Rate) May 2025'] = df['Updated DP/Mtr (Cut Rate) May 2025'].apply(clean_numeric_value)
    
    if 'GST' in df.columns:
        df['GST'] = df['GST'].apply(clean_numeric_value)
        # Convert percentage to decimal if needed (e.g., 5% -> 0.05)
        df['GST'] = df['GST'].apply(lambda x: x/100 if x and x > 1 else x)
    
    # Calculate Dealer Price after GST if missing
    if 'Dealer Price after GST (Cut Rate)' in df.columns:
        mask = df['Dealer Price after GST (Cut Rate)'].isna()
        if mask.any() and 'Updated DP/Mtr (Cut Rate) May 2025' in df.columns and 'GST' in df.columns:
            df.loc[mask, 'Dealer Price after GST (Cut Rate)'] = (
                df.loc[mask, 'Updated DP/Mtr (Cut Rate) May 2025'] * (1 + df.loc[mask, 'GST'].fillna(0))
            )
    
    # Calculate RR Price after GST if missing
    if 'RR Price after GST (Cut Rate)' in df.columns and 'RR Price (Cut Rate)' in df.columns:
        mask = df['RR Price after GST (Cut Rate)'].isna()
        if mask.any() and 'GST' in df.columns:
            df.loc[mask, 'RR Price after GST (Cut Rate)'] = (
                df.loc[mask, 'RR Price (Cut Rate)'] * (1 + df.loc[mask, 'GST'].fillna(0))
            )
    
    return df

def apply_vaya_formatting(workbook, sheet):
    """Apply VAYA-style formatting to the sheet"""
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='center')
    number_alignment = Alignment(horizontal='right', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Format header row
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Set row height for header
    sheet.row_dimensions[1].height = 30
    
    # Format data rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for idx, cell in enumerate(row):
            cell.font = cell_font
            cell.border = border
            
            # Right-align numeric columns
            if idx in [3, 4, 5, 6, 7, 10, 11, 12, 13, 14, 15, 16]:  # Numeric columns
                cell.alignment = number_alignment
            else:
                cell.alignment = cell_alignment
    
    # Set column widths
    column_widths = {
        'A': 25,  # Collection
        'B': 20,  # Design Name
        'C': 20,  # Composition
        'D': 12,  # Weight/Mt
        'E': 15,  # Design Horizontal
        'F': 15,  # Design Vertical
        'G': 15,  # Fabric Width
        'H': 12,  # Martindale
        'I': 15,  # End Use
        'J': 12,  # HS Code
        'K': 20,  # Updated DP
        'L': 8,   # GST
        'M': 20,  # Dealer Price after GST (Cut)
        'N': 20,  # Dealer Price (Roll)
        'O': 20,  # Dealer Price after GST (Roll)
        'P': 20,  # RR Price (Cut)
        'Q': 25   # RR Price after GST (Cut)
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
    
    # Freeze header row
    sheet.freeze_panes = 'A2'

def standardize_catalog(input_file, output_file=None):
    """Main function to standardize a catalog to VAYA format"""
    
    input_path = Path(input_file)
    
    if not input_path.exists():
        print(f"Error: File not found: {input_file}")
        return False
    
    # Generate output filename if not provided
    if output_file is None:
        output_file = input_path.parent / f"{input_path.stem}_STANDARDIZED.xlsx"
    
    try:
        # Read the Excel file
        print(f"Reading {input_path.name}...")
        
        # Try to read all sheets
        all_sheets = pd.read_excel(input_file, sheet_name=None)
        
        # Create output workbook
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)  # Remove default sheet
        
        sheet_count = 0
        for sheet_name, df in all_sheets.items():
            if df.empty:
                continue
            
            print(f"\nProcessing sheet: {sheet_name}")
            
            # Detect catalog type
            catalog_type = detect_catalog_type(df, input_path.name)
            print(f"  Detected type: {catalog_type}")
            
            # Map columns to VAYA standard
            standardized_df = map_columns(df, catalog_type)
            
            # Calculate GST prices
            standardized_df = calculate_gst_prices(standardized_df)
            
            # Remove completely empty rows
            standardized_df = standardized_df.dropna(how='all')
            
            # Create sheet
            ws = wb_out.create_sheet(title=sheet_name)
            
            # Write headers
            for col_idx, col_name in enumerate(VAYA_COLUMNS, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write data
            for row_idx, row in enumerate(standardized_df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply formatting
            apply_vaya_formatting(wb_out, ws)
            
            sheet_count += 1
        
        # Save workbook
        wb_out.save(output_file)
        print(f"\n✅ SUCCESS! Standardized catalog saved to: {output_file}")
        print(f"   Processed {sheet_count} sheet(s)")
        
        return True
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python catalog_standardizer.py <input_file> [output_file]")
        print("\nExample:")
        print("  python catalog_standardizer.py SANSAAR_NEW_PRICE_LIST.xlsx")
        print("  python catalog_standardizer.py input.xlsx output_standardized.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    success = standardize_catalog(input_file, output_file)
    sys.exit(0 if success else 1)